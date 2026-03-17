from django.contrib import admin, messages
from django.utils import timezone
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.admin import UserAdmin as DjangoUserAdmin
from unfold.admin import ModelAdmin
from django.contrib.auth.forms import UserCreationForm, UserChangeForm
from .models import User, Branch, Department
from django.http import HttpResponseRedirect
from django.urls import reverse
from .utils.user_import import make_base_username, pick_username, generate_password_8
from openpyxl import load_workbook, Workbook
from django.urls import path
from exams.models import TestRule
from io import BytesIO
from django.db import transaction
from django.db.models import Q
from django.utils.safestring import mark_safe
from django.utils.html import format_html

@admin.register(Branch)
class BranchAdmin(ModelAdmin):
    list_display = ("id", "name")
    search_fields = ("name",)
    list_display_links = ("name",)
    list_per_page = 20              # default 20 ta
    list_max_show_all = 200         # "Show all" limiti
    list_per_page_options = [20, 50, 100]

@admin.register(Department)
class DepartmentAdmin(ModelAdmin):
    list_display = ("id", "name",)
    search_fields = ("name",)
    list_display_links = ("name",)
    list_per_page = 20              # default 20 ta
    list_max_show_all = 200         # "Show all" limiti
    list_per_page_options = [20, 50, 100]

class UserCreateForm(UserCreationForm):
    class Meta(UserCreationForm.Meta):
        model = User
        fields = (
            "username", "first_name", "last_name", "patronymic",
            "branch", "department",
            "role", "is_active", "phone", "email",
        )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        # Uzbekcha label'lar
        self.fields["username"].label = "Login"
        self.fields["first_name"].label = "Ismi"
        self.fields["last_name"].label = "Familiyasi"
        self.fields["patronymic"].label = "Otasining ismi"
        self.fields["branch"].label = "Filial"
        self.fields["department"].label = "Bo‘lim"
        self.fields["role"].label = "Rol"
        self.fields["is_active"].label = "Aktiv"
        self.fields["phone"].label = "Telefon raqami"
        self.fields["email"].label = "Email"

        # Faqat YANGI user yaratishda majburiy
        self.fields["first_name"].required = True
        self.fields["last_name"].required = True
        self.fields["department"].required = True
        self.fields["branch"].required = True
        self.fields["department"].required = True
    
    def clean(self):
        cleaned_data = super().clean()

        if not cleaned_data.get("phone"):
            cleaned_data["phone"] = None

        if not cleaned_data.get("email"):
            cleaned_data["email"] = None

        return cleaned_data
class UserEditForm(UserChangeForm):
    class Meta(UserChangeForm.Meta):
        model = User
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        # Uzbekcha label'lar
        if "username" in self.fields:
            self.fields["username"].label = "Login"
        if "first_name" in self.fields:
            self.fields["first_name"].label = "Ismi"
        if "last_name" in self.fields:
            self.fields["last_name"].label = "Familiyasi"
        if "patronymic" in self.fields:
            self.fields["patronymic"].label = "Otasining ismi"
        if "branch" in self.fields:
            self.fields["branch"].label = "Filial"
        if "department" in self.fields:
            self.fields["department"].label = "Bo‘lim"
        if "role" in self.fields:
            self.fields["role"].label = "Rol"
        if "is_active" in self.fields:
            self.fields["is_active"].label = "Aktiv"
        if "phone" in self.fields:
            self.fields["phone"].label = "Telefon raqami"
        if "email" in self.fields:
            self.fields["email"].label = "Email"



def _chip(text: str) -> str:
    return (
        f'<span class="px-2 py-0.5 rounded-full text-[11px] '
        f'bg-slate-700/50 text-slate-200">'
        f'{text}</span>'
    )



##########################USERADMIN#########################################
@admin.register(User)
class UserAdmin(ModelAdmin, DjangoUserAdmin):
    add_form = UserCreateForm
    form = UserEditForm
    list_per_page = 50          # default 50
    list_max_show_all = 200 
    class Media:
        css = {
            "all": ("admin/users_changelist.css",)
        }
    list_display = (
        "id",
        "username",
        "full_name", 
        "branch",
        "department",
        "tests_chips",
        "role",
        "is_active",
        "phone",
        "email",
    )
    def full_name(self, obj):
        return format_html(
            """
            <div style="font-size:12px;">
                <div style="font-weight:600;">{} {}</div>
                <div style="font-size:11px; opacity:0.6;">{}</div>
            </div>
            """,
            obj.last_name or "",
            obj.first_name or "",
            obj.patronymic or "",
        )

    full_name.short_description = "F.I.O"
    list_display_links = ("username",)
    filter_horizontal = ("assigned_tests",)
    list_filter = ("branch", "department", "role", "is_active")
    search_fields = ("username", "last_name", "first_name",  "phone", "email")
    ordering = ("id",)
    
    def tests_chips(self, obj):
        # 1) Direct (userga qo‘lda biriktirilgan)
        direct = list(obj.assigned_tests.all())

        # 2) Rule bo‘yicha ruxsat berilgan
        rules = TestRule.objects.filter(is_active=True, test__is_active=True)

        # filial qoidasi: None (global) yoki user.branch ga teng
        rules = rules.filter(Q(branch__isnull=True) | Q(branch=obj.branch))

        # bo‘lim qoidasi: None (global) yoki user.department ga teng
        rules = rules.filter(Q(department__isnull=True) | Q(department=obj.department))

        # rol qoidasi: None/bo‘sh (global) yoki user.role ga teng
        rules = rules.filter(Q(role__isnull=True) | Q(role="") | Q(role=obj.role))

        rule_tests = [r.test for r in rules.select_related("test")]

        # 3) Birlashtiramiz (dublikatlarsiz)
        all_tests = {}
        for t in direct + rule_tests:
            all_tests[t.id] = t

        if not all_tests:
            return "-"

        chips_html = " ".join(_chip(t.title) for t in all_tests.values())
        return mark_safe(chips_html)


    tests_chips.short_description = "Biriktirilgan testlar"


    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.select_related("department").prefetch_related(
            "assigned_tests",
        )


    change_list_template = "admin/users/user/change_list.html"
    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path("import-xlsx/", self.admin_site.admin_view(self.import_xlsx), name="users_user_import_xlsx"),
            path("import-xlsx/sample/", self.admin_site.admin_view(self.import_xlsx_sample), name="users_user_import_xlsx_sample"),
        ]
        return my_urls + urls


    def import_xlsx(self, request):
        if request.method == "GET":
            context = dict(self.admin_site.each_context(request), title="Users import (XLSX)")
            return render(request, "admin/users/user/import_xlsx.html", context)

        file = request.FILES.get("file")
        if not file:
            messages.error(request, "XLSX fayl tanlanmadi.")
            context = dict(self.admin_site.each_context(request), title="Users import (XLSX)")
            return render(request, "admin/users/user/import_xlsx.html", context)


        wb = load_workbook(file)
        ws = wb.active

        # headerlarni normalize qilamiz (Bo‘lim/Bo'lim, Telefon/telefon, Email/email)
        header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

        def find_col(*names):
            names_norm = {n.strip().lower().replace("‘", "'") for n in names}
            for idx, h in enumerate(header):
                hn = h.strip().lower().replace("‘", "'")
                if hn in names_norm:
                    return idx
            return None

        c_first = find_col("Ismi", "ismi")
        c_last = find_col("Familiyasi", "familiyasi")
        c_pat = find_col("Otasining ismi", "otasining ismi", "Otasining ismi ")
        c_branch = find_col("Filial", "filial")
        c_dept = find_col("Bo'lim", "Bo‘lim", "bo'lim", "bo‘lim")
        c_phone = find_col("Telefon", "telefon")
        c_email = find_col("Email", "email")

        for name, c in [("Ismi", c_first), ("Familiyasi", c_last), ("Otasining ismi", c_pat), ("Filial", c_branch), ("Bo'lim", c_dept)]:
            if c is None:
                messages.error(request, f"Ustun topilmadi: {name}")
                return render(request, "admin/users/user/import_xlsx.html", {})

        # Natija export (importdan keyin darhol)
        out = Workbook()
        out_ws = out.active
        out_ws.title = "Natija"
        out_ws.append([
            "username", "password",
            "Ismi", "Familiyasi", "Otasining ismi",
            "Filial", "Bo'lim", "Telefon", "Email",
            "status"
        ])

        created = 0
        exists = 0
        errors = 0

        # MUHIM: Sizda "Filial" modeli qayerda?
        # Screenshotda Department bor. Agar Department ichida filial/branch field bo'lsa, shundan foydalanamiz.
        # Hozircha branch_name ni Department orqali tekshirish uchun quyidagi kabi qilamiz:
        # - Department: name=bo'lim, (optional) branch/filial = branch_name

        for row in ws.iter_rows(min_row=2, values_only=True):
            first_name = (row[c_first] or "").strip()
            last_name = (row[c_last] or "").strip()
            patronymic = (row[c_pat] or "").strip()
            branch_name = (row[c_branch] or "").strip()
            dept_name = (row[c_dept] or "").strip()

            phone = None
            if c_phone is not None:
                p = str(row[c_phone] or "").strip()
                phone = p if p else None

            email = None
            if c_email is not None:
                e = str(row[c_email] or "").strip()
                email = e if e else None

            try:
                # majburiy tekshiruv
                if not (first_name and last_name and branch_name and dept_name):
                    raise ValueError("Majburiy maydonlar to‘liq emas (Ismi/Familiyasi/Filial/Bo'lim).")

                # 0) Filial topish
                branch = Branch.objects.filter(name__iexact=branch_name).first()
                if not branch:
                    raise ValueError(f"Filial topilmadi: {branch_name}")

                # 1) Bo‘lim topish (filialga bog‘liq EMAS)
                dept = Department.objects.filter(name__iexact=dept_name).first()
                if not dept:
                    raise ValueError(f"Bo'lim topilmadi: {dept_name}")

                base = make_base_username(first_name, patronymic, last_name)

                # ✅ EXISTS sharti: bir xil filial+bo'lim ichida base'dan boshlangan username bo'lsa (a.b, a.b12, a.b34)
                if User.objects.filter(branch=branch, department=dept, username__startswith=base).exists():
                    exists += 1
                    out_ws.append([
                        base, "",
                        first_name, last_name, patronymic,
                        branch_name, dept_name,
                        phone or "", email or "",
                        "EXISTS"
                    ])
                    continue

                # ✅ username generatsiya (unik)
                username = pick_username(User, base)
                password = generate_password_8()

                u = User(
                    username=username,
                    first_name=first_name,
                    last_name=last_name,
                    patronymic=patronymic,
                    branch=branch,
                    department=dept,
                    phone=phone,
                    email=email,
                )
                u.set_password(password)
                u.save()

                created += 1
                out_ws.append([
                    username, password,
                    first_name, last_name, patronymic,
                    branch_name, dept_name,
                    phone or "", email or "",
                    "CREATED"
                ])

            except Exception as ex:
                errors += 1
                out_ws.append([
                    "", "",
                    first_name, last_name, patronymic,
                    branch_name, dept_name,
                    phone or "", email or "",
                    f"ERROR: {ex}"
                ])

        filename = f"import_result_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        buffer = BytesIO()
        out.save(buffer)
        buffer.seek(0)

        resp = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp


    # Add qilayotganda ko‘rinadigan maydonlar
    add_fieldsets = (
        (None, {
            "classes": ("wide",),
            "fields": (
                "username",
                "password1",
                "password2",
                "first_name",
                "last_name",
                "patronymic",
                "branch",
                "department",
                "assigned_tests",
                "role",
                "is_active",
                "phone",
                "email",
            ),
        }),
    )

    # Edit sahifasi bo‘limlari
    fieldsets = (
        (None, {"fields": ("username", "password")}),
        ("Shaxsiy", {"fields": ("first_name", "last_name", "patronymic")}),
        ("Ish joyi", {"fields": ("branch", "department")}),
        ("Testlar", {"fields": ("assigned_tests",)}),
        ("Aloqa", {"fields": ("phone", "email")}),
        ("Holat va rol", {"fields": ("role", "is_active")}),
        ("Ruxsatlar", {"fields": ("is_staff", "is_superuser", "groups", "user_permissions")}),
        ("Sanalar", {"fields": ("last_login", "date_joined")}),
    )
    def response_add(self, request, obj, post_url_continue=None):
        """
        User saqlangandan keyin doim Add sahifaga qaytadi (bo'sh forma).
        Ya'ni "Save" ham "Save and add another" kabi ishlaydi.
        """
        # Django default response'ni chaqiramiz (message va loglar uchun)
        super().response_add(request, obj, post_url_continue=post_url_continue)

        # Har doim yangi bo'sh add formga qaytamiz (querysiz)
        add_url = reverse("admin:users_user_add")
        return HttpResponseRedirect(add_url)
    
    def import_xlsx_sample(self, request):
        from openpyxl import Workbook
        from django.http import HttpResponse

        wb = Workbook()
        ws = wb.active
        ws.title = "Sample"

        # Siz so‘ragan headerlar
        ws.append(["Ismi", "Familiyasi", "Otasining ismi", "Filial", "Bo'lim", "Telefon", "Email"])

        # 2 ta namunaviy qator
        ws.append(["Abulqosim", "Begaliyev", "To'lqin o'g'li", "Boshqaruv apparati", "Raqamlashtirish boshqarmasi", "+998901234567", "test1@example.com"])
        ws.append(["Aziza", "Karimova", "Anvar qizi", "Boshqaruv apparati", "Raqamlashtirish boshqarmasi", "", ""])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        resp = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="users_import_sample.xlsx"'
        return resp

    actions = ["export_xlsx"]

    def export_xlsx(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"

        ws.append([
            "username",
            "Ismi",
            "Familiyasi",
            "Otasining ismi",
            "Filial",
            "Bo'lim",
            "Telefon",
            "Email",
            "Rol",
            "Aktiv",
            "Testlar (bo'lim+user)",
        ])

        # tezroq bo‘lishi uchun
        queryset = queryset.select_related("branch", "department").prefetch_related(
            "assigned_tests",
        )

        for u in queryset:
            dept_tests = list(u.department.assigned_tests.values_list("title", flat=True)) if u.department_id else []
            user_tests = list(u.assigned_tests.values_list("title", flat=True))
            tests = sorted(set(dept_tests + user_tests))

            ws.append([
                u.username,
                u.first_name,
                u.last_name,
                getattr(u, "patronymic", ""),
                u.branch.name if u.branch_id else "",
                u.department.name if u.department_id else "",
                u.phone or "",
                u.email or "",
                u.role,
                "Ha" if u.is_active else "Yo‘q",
                ", ".join(tests),
            ])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"users_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        resp = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp

    export_xlsx.short_description = "Export XLSX (tanlanganlar)"




    actions = ["reset_passwords_export_xlsx"]

    @admin.action(description="Parolni qayta tiklash (XLSX)")
    def reset_passwords_export_xlsx(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Reset passwords"

        ws.append([
            "username", "new_password",
            "Ismi", "Familiyasi", "Otasining ismi",
            "Filial", "Bo'lim",
            "Telefon", "Email",
            "status",
        ])

        queryset = queryset.select_related("branch", "department").order_by("id")

        # hammasi bir tranzaksiyada (xohlasangiz olib tashlasa ham bo'ladi)
        with transaction.atomic():
            for u in queryset:
                try:
                    new_pass = generate_password_8()
                    u.set_password(new_pass)
                    u.save(update_fields=["password"])

                    ws.append([
                        u.username, new_pass,
                        u.first_name, u.last_name, getattr(u, "patronymic", ""),
                        u.branch.name if u.branch_id else "",
                        u.department.name if u.department_id else "",
                        u.phone or "",
                        u.email or "",
                        "RESET",
                    ])
                except Exception as ex:
                    ws.append([
                        u.username, "",
                        u.first_name, u.last_name, getattr(u, "patronymic", ""),
                        u.branch.name if u.branch_id else "",
                        u.department.name if u.department_id else "",
                        u.phone or "",
                        u.email or "",
                        f"ERROR: {ex}",
                    ])

        filename = f"password_reset_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        resp = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp
