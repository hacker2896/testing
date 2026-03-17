from django.contrib import admin
from unfold.admin import ModelAdmin  # <-- MUHIM
from .models import Test, Question, Choice, Attempt, AttemptAnswer, TestRule
from django.core.exceptions import ValidationError
from django import forms
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from django.shortcuts import render, redirect
from openpyxl import load_workbook, Workbook
from django.contrib import admin, messages
from django.db import transaction
from django.urls import path
import re
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.http import HttpResponse
from .models import AttemptPolicy

def _norm(s: str) -> str:
        return " ".join((s or "").strip().lower().split())


def _cell_text(row, idx: int) -> str:
    if idx >= len(row):
        return ""
    v = row[idx].value
    return "" if v is None else str(v).strip()

    
def _parse_correct_numbers(raw: str) -> set[int]:
        """
    Correct ustunidan raqamlarni universal ajratadi.
    Qabul qiladi:
            1
            1;3
            1,3
            1 3
            1|3
        """
        raw = (raw or "").strip()
        if not raw:
            return set()

        # har qanday ajratuvchini bo'sh joyga almashtiramiz
        parts = re.split(r"[^\d]+", raw)

        return {int(p) for p in parts if p.isdigit()}
    

    




class ImportQuestionsXlsxForm(forms.Form):
    file = forms.FileField(label="XLSX fayl")
    overwrite = forms.BooleanField(
        required=False,
        initial=False,
        label="Savol mavjud bo'lsa (test+savol bo'yicha) yangilansin"
    )






class ChoiceInlineFormSet(forms.BaseInlineFormSet):
    def clean(self):
        super().clean()

        correct = 0
        total = 0

        for form in self.forms:
            # o‘chirilayotgan qatorlarni hisobga olmaymiz
            if not hasattr(form, "cleaned_data"):
                continue
            if form.cleaned_data.get("DELETE"):
                continue

            total += 1
            if form.cleaned_data.get("is_correct"):
                correct += 1

        # kamida 1 ta variant bo'lsin (xohlasangiz)
        if total == 0:
            raise ValidationError("Kamida bitta javob varianti kiritilishi kerak.")

        # kamida 1 ta to'g'ri javob bo'lsin
        if correct < 1:
            raise ValidationError("Kamida bitta 'To‘g‘ri javob' belgilanishi shart.")

        # single choice bo'lsa: faqat 1 ta to'g'ri
        if correct > 1:
            raise ValidationError("Faqat bitta 'To‘g‘ri javob' belgilash mumkin.")
        

class ChoiceInline(admin.TabularInline):
    model = Choice
    extra = 0 
    validate_min = True
    formset = ChoiceInlineFormSet
    fields = ("text", "is_correct")
    autocomplete_fields = ()
    # Unfold ko'rinishini chiroyli qilish uchun
    classes = ("collapse",)  # xohlasangiz olib tashlang

@admin.register(Question)
class QuestionAdmin(ModelAdmin):  # <-- admin.ModelAdmin emas
    list_display = ("id", "small_text", "question_type","points", "test", )
    list_filter = ("test", "question_type",)
    search_fields = ("small_text",)
    list_display_links = ("small_text",)
    inlines = [ChoiceInline]
    list_per_page = 20              # default 20 ta
    list_max_show_all = 200         # "Show all" limiti
    list_per_page_options = [20, 50, 100]
    change_list_template = "admin/exams/question/change_list.html"
    actions = ["export_selected_xlsx", "export_filtered_xlsx"]


    def get_inline_instances(self, request, obj=None):
        if obj and obj.question_type in {"short", "numeric", "essay"}:
            return []
        return super().get_inline_instances(request, obj)

    # true_false bo'lsa, saqlaganda 2ta variantni avtomatik yaratib qo'yamiz (yo'q bo'lsa)
    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)
        q = form.instance

        if q.question_type == "true_false":
            qs = q.choices.all()  # related_name="choices" bo'lsa; bo'lmasa choice_set
            if not qs.exists():
                Choice.objects.create(question=q, text="True", is_correct=True)
                Choice.objects.create(question=q, text="False", is_correct=False)

    class Media:
        js = ("admin/question_type_ui.js",)

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path("import-xlsx/", self.admin_site.admin_view(self.import_xlsx), name="exams_question_import_xlsx"),
            path("import-template/", self.admin_site.admin_view(self.download_template), name="exams_question_import_template"),
        ]
        return custom + urls
    
    def download_template(self, request):
        """
        XLSX template'ni dinamik generatsiya qiladi (diskdan qidirmaydi).
        Format: Test | Type | Question | Correct | Answer1..Answer6
        Correct: single=1, multiple=1;3, true_false=1/2, short/numeric=matn/son, essay=bo'sh.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "questions"

        headers = ["Test", "Type", "Question", "Correct",
                "Answer1", "Answer2", "Answer3", "Answer4", "Answer5", "Answer6"]
        ws.append(headers)

        sample_rows = [
            # SINGLE (1 ta to'g'ri)
            ["Kompyuter savodxonligi", "single",
            "Monitor nima?", "1",
            "Qurilma", "Dastur", "Fayl", "", "", ""],

            # MULTIPLE (bir nechta to'g'ri) - Correct raqamlar; ; yoki , ishlaydi
            ["Kompyuter savodxonligi", "multiple",
            "Qaysilari kiritish qurilmalari?", "2;4",
            "Printer", "Klaviatura", "Monitor", "Sichqoncha", "", ""],

            # TRUE/FALSE (1=True, 2=False)
            ["Kompyuter savodxonligi", "true_false",
            "Kompyuter elektron qurilma.", "1",
            "", "", "", "", "", ""],

            # SHORT (correct_answer ga yoziladi)
            ["Kompyuter savodxonligi", "short",
            "CPU nimani anglatadi?", "Central Processing Unit",
            "", "", "", "", "", ""],

            # NUMERIC
            ["Kompyuter savodxonligi", "numeric",
            "2 + 2 nechiga teng?", "4",
            "", "", "", "", "", ""],

            # ESSAY (Correct bo'sh bo'lishi mumkin)
            ["Kompyuter savodxonligi", "essay",
            "Kompyuter xavfsizligi bo‘yicha qisqa izoh bering.", "",
            "", "", "", "", "", ""],
        ]

        for r in sample_rows:
            ws.append(r)

        # ustun kengliklari
        widths = [26, 12, 60, 18, 22, 22, 22, 22, 22, 22]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        resp = HttpResponse(
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = 'attachment; filename="questions_import_template.xlsx"'
        return resp
    

    def import_xlsx(self, request):
        """
        Excel format (pozitsiya bo'yicha):
        0 Test
        1 Type (single/multiple/true_false/short/numeric/essay)
        2 Question
        3 Correct
        4..N Answers

        Test topilmasa -> SKIP (xato sifatida).
        Correct raqamli:
        single: 1
        multiple: 1;3;4
        true_false: 1(True) yoki 2(False)
        short/numeric: matn/son
        essay: bo'sh mumkin
        """
        # --- Sizda yuqorida form bo'lishi kerak:
        # class ImportQuestionsXlsxForm(forms.Form):
        #   file = forms.FileField(...)
        #   overwrite = forms.BooleanField(...)

        from .models import Test, Question, Choice  # <-- kerak bo'lsa moslang
        from .admin import ImportQuestionsXlsxForm  # <-- agar form shu faylda bo'lsa o'zgartiring / olib tashlang

        allowed_types = {"single", "multiple", "true_false", "short", "numeric", "essay"}

        # --- Model fieldlarini "avto topish"
        test_name_field = "title" if hasattr(Test, "title") else ("name" if hasattr(Test, "name") else None)
        question_text_field = "text" if hasattr(Question, "text") else ("question" if hasattr(Question, "question") else None)
        choice_text_field = "text" if hasattr(Choice, "text") else ("title" if hasattr(Choice, "title") else None)
        choice_correct_field = "is_correct" if hasattr(Choice, "is_correct") else ("correct" if hasattr(Choice, "correct") else None)

        if not test_name_field or not question_text_field or not choice_text_field or not choice_correct_field:
            messages.error(
                request,
                "Model field nomlarini aniqlab bo'lmadi. "
                "Test(title/name), Question(text/question), Choice(text, is_correct) bo'lishi kerak."
            )
            return redirect("..")

        if request.method == "POST":
            form = ImportQuestionsXlsxForm(request.POST, request.FILES)
            if form.is_valid():
                f = form.cleaned_data["file"]
                overwrite = form.cleaned_data["overwrite"]

                wb = load_workbook(f, data_only=True)
                ws = wb.active

                rows = list(ws.iter_rows())
                if not rows:
                    messages.error(request, "XLSX bo'sh.")
                    return redirect("..")

                # Header bor-yo'qligini taxmin qilish (ixtiyoriy)
                first4 = [_norm(_cell_text(rows[0], i)) for i in range(4)]
                looks_like_header = any(x in {"test", "type", "question", "correct"} for x in first4)
                start_i = 1 if looks_like_header else 0

                created_q = updated_q = 0
                created_c = 0
                skipped = 0
                errors: list[tuple[int, str]] = []

                def get_test_by_name(name: str):
                    # exact match; xohlasangiz icontains ham qo'shamiz
                    return Test.objects.filter(**{test_name_field: name}).first()

                with transaction.atomic():
                    for r_index in range(start_i, len(rows)):
                        row = rows[r_index]
                        excel_row_no = r_index + 1

                        test_name = _cell_text(row, 0)
                        q_type = _norm(_cell_text(row, 1))
                        q_text = _cell_text(row, 2)
                        correct_raw = _cell_text(row, 3)

                        if not test_name and not q_text:
                            continue  # bo'sh qator

                        if not test_name or not q_text:
                            skipped += 1
                            errors.append((excel_row_no, "Test nomi yoki Savol bo'sh"))
                            continue

                        if not q_type:
                            q_type = "single"
                        if q_type not in allowed_types:
                            skipped += 1
                            errors.append((excel_row_no, f"Type noto'g'ri: {q_type}"))
                            continue

                        test = get_test_by_name(test_name)
                        if not test:
                            skipped += 1
                            errors.append((excel_row_no, f"Test topilmadi: {test_name}"))
                            continue

                        # answers (4..N)
                        answers: list[str] = []
                        for j in range(4, len(row)):
                            a = _cell_text(row, j)
                            if a:
                                answers.append(a)

                        # overwrite bo'lsa savolni topamiz
                        q_obj = None
                        if overwrite:
                            q_obj = Question.objects.filter(
                                test=test,
                                **{question_text_field: q_text},
                            ).first()

                        # create/update savol
                        if q_obj:
                            q_obj.question_type = q_type
                            # short/numeric uchun keyin to'ldiramiz, hozir tozalaymiz
                            if hasattr(q_obj, "correct_answer"):
                                q_obj.correct_answer = None
                            q_obj.save()
                            updated_q += 1

                            # eski variantlarni tozalash
                            Choice.objects.filter(question=q_obj).delete()
                        else:
                            q_obj = Question.objects.create(
                                test=test,
                                **{question_text_field: q_text},
                                question_type=q_type,
                            )
                            created_q += 1

                        # TYPE bo'yicha ishlov
                        if q_type in {"single", "multiple"}:
                            if len(answers) < 2:
                                skipped += 1
                                errors.append((excel_row_no, "Variantlar 2 tadan kam"))
                                q_obj.delete()
                                continue

                            correct_set = _parse_correct_numbers(correct_raw)

                            if q_type == "single":
                                if len(correct_set) != 1:
                                    skipped += 1
                                    errors.append((excel_row_no, "single uchun Correct faqat bitta raqam bo'lishi kerak (masalan: 2)"))
                                    q_obj.delete()
                                    continue
                            else:
                                # multiple
                                if len(correct_set) < 1:
                                    skipped += 1
                                    errors.append((excel_row_no, "multiple uchun Correct kamida 1 ta bo'lishi kerak (masalan: 1;3)"))
                                    q_obj.delete()
                                    continue

                            # Correct raqamlari diapazonda bo'lsin
                            if any(n < 1 or n > len(answers) for n in correct_set):
                                skipped += 1
                                errors.append((excel_row_no, f"Correct raqami diapazondan tashqarida. Variantlar soni={len(answers)}, Correct={correct_raw}"))
                                q_obj.delete()
                                continue

                            found = 0
                            for idx, a in enumerate(answers, start=1):
                                is_ok = idx in correct_set
                                if is_ok:
                                    found += 1
                                Choice.objects.create(
                                    question=q_obj,
                                    **{choice_text_field: a},
                                    **{choice_correct_field: is_ok},
                                )
                                created_c += 1

                            if q_type == "single" and found != 1:
                                skipped += 1
                                errors.append((excel_row_no, f"single uchun 1 ta correct chiqishi kerak edi, topildi: {found}"))
                                q_obj.delete()
                                continue
                            if q_type == "multiple" and found < 1:
                                skipped += 1
                                errors.append((excel_row_no, "multiple uchun hech bo'lmasa 1 ta correct topilishi kerak"))
                                q_obj.delete()
                                continue

                        elif q_type == "true_false":
                            # Correct: 1=True, 2=False
                            correct_set = _parse_correct_numbers(correct_raw)
                            if correct_set not in ({1}, {2}):
                                skipped += 1
                                errors.append((excel_row_no, "true_false uchun Correct 1 (True) yoki 2 (False) bo'lishi kerak"))
                                q_obj.delete()
                                continue

                            Choice.objects.create(
                                question=q_obj,
                                **{choice_text_field: "True"},
                                **{choice_correct_field: (1 in correct_set)},
                            )
                            Choice.objects.create(
                                question=q_obj,
                                **{choice_text_field: "False"},
                                **{choice_correct_field: (2 in correct_set)},
                            )
                            created_c += 2

                        elif q_type in {"short", "numeric"}:
                            if not correct_raw:
                                skipped += 1
                                errors.append((excel_row_no, f"{q_type} uchun Correct bo'sh bo'lmasligi kerak"))
                                q_obj.delete()
                                continue

                            if not hasattr(q_obj, "correct_answer"):
                                skipped += 1
                                errors.append((excel_row_no, f"{q_type} uchun Question modelda correct_answer field yo'q"))
                                q_obj.delete()
                                continue

                            q_obj.correct_answer = correct_raw
                            q_obj.save()

                        elif q_type == "essay":
                            # essay uchun correct ixtiyoriy
                            if hasattr(q_obj, "correct_answer"):
                                q_obj.correct_answer = correct_raw or None
                                q_obj.save()

                messages.success(
                    request,
                    f"Import tugadi. Savol: +{created_q}, yangilandi: {updated_q}, Variantlar: +{created_c}, Skipped: {skipped}."
                )
                if errors:
                    preview = "\n".join([f"Row {r}: {reason}" for r, reason in errors[:10]])
                    messages.warning(request, f"Birinchi xatolar:\n{preview}")

                return redirect("..")

        else:
            form = ImportQuestionsXlsxForm()

        context = dict(
            self.admin_site.each_context(request),
            title="Savollarni XLSX dan import qilish",
            form=form,
            # Import sahifasida "Template yuklab olish" linki ko'rinsin desangiz template'da ishlating:
            # template_download_url = "../import-template/"  (siz get_urls()da alohida view qo'shasiz)
            template_download_url="../import-template/",
        )
        return render(request, "admin/exams/question/import_xlsx.html", context)

    #########export all selected###############
    def build_xlsx(queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "questions"

        headers = [
            "ID", "Test", "Type", "Question",
            "Correct", "Answer1", "Answer2",
            "Answer3", "Answer4", "Answer5", "Answer6"
        ]
        ws.append(headers)

        for q in queryset.select_related("test"):
            qtype = q.question_type
            qtext = q.text

            # choices
            choices = list(q.choices.all()) if hasattr(q, "choices") else list(q.choice_set.all())
            answers = [c.text for c in choices][:6]
            answers += [""] * (6 - len(answers))

            # correct
            correct = ""
            if qtype in ("single", "multiple", "true_false"):
                idxs = [str(i + 1) for i, c in enumerate(choices) if c.is_correct]
                if qtype in ("single", "true_false"):
                    correct = idxs[0] if idxs else ""
                else:
                    correct = ";".join(idxs)
            else:
                correct = getattr(q, "correct_answer", "") or ""

            ws.append([
                q.id,
                q.test.title if q.test else "",
                qtype,
                qtext,
                correct,
                *answers
            ])

        # column widths
        widths = [8, 28, 14, 90, 16, 24, 24, 24, 24, 24, 24]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"
        return wb

#####export###########

    def build_xlsx(self, queryset):
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "questions"

        headers = [
            "ID", "Test", "Type", "Question",
            "Correct", "Answer1", "Answer2",
            "Answer3", "Answer4", "Answer5", "Answer6"
        ]
        ws.append(headers)

        for q in queryset.select_related("test"):
            qtype = q.question_type
            qtext = q.text

            choices = list(q.choices.all()) if hasattr(q, "choices") else list(q.choice_set.all())
            answers = [c.text for c in choices][:6]
            answers += [""] * (6 - len(answers))

            correct = ""
            if qtype in ("single", "multiple", "true_false"):
                idxs = [str(i + 1) for i, c in enumerate(choices) if c.is_correct]
                if qtype in ("single", "true_false"):
                    correct = idxs[0] if idxs else ""
                else:
                    correct = ";".join(idxs)
            else:
                correct = getattr(q, "correct_answer", "") or ""

            ws.append([
                q.id,
                q.test.title if q.test else "",
                qtype,
                qtext,
                correct,
                *answers
            ])

        widths = [8, 28, 14, 90, 16, 24, 24, 24, 24, 24, 24]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"
        return wb


    @admin.action(description="📤 Export XLSX (tanlangan)")
    def export_selected_xlsx(self, request, queryset):
        wb = self.build_xlsx(queryset)

        from django.http import HttpResponse
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="questions_selected.xlsx"'
        wb.save(response)
        return response


    @admin.action(description="📤 Export XLSX (filter bo‘yicha hammasi)")
    def export_filtered_xlsx(self, request, queryset):
        changelist = self.get_changelist_instance(request)
        filtered_qs = changelist.get_queryset(request)

        wb = self.build_xlsx(filtered_qs)

        from django.http import HttpResponse
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="questions_filtered.xlsx"'
        wb.save(response)
        return response



    def small_text(self, obj):
        return format_html(
            '''
            <div style="
                font-size: 13px;
                line-height: 1.4;
                max-width: 600px;
                display: -webkit-box;
                -webkit-line-clamp: 3;
                -webkit-box-orient: vertical;
                overflow: hidden;
            ">
                {}
            </div>
            ''',
            obj.text
        )

    small_text.short_description = "Savol matni"

class TestRuleInline(admin.TabularInline):
    model = TestRule
    extra = 1

@admin.register(Test)
class TestAdmin(ModelAdmin):  # <-- admin.ModelAdmin emas
    list_display = ("id", "title", "code", "questions_count", "questions_total", "pass_percent", "is_active")
    search_fields = ("title", "code")
    list_filter = ("is_active","title",)
    list_display_links = ("title",)
    def questions_total(self, obj):
        return obj.questions.count()
    questions_total.short_description = "Bankdagi savollar"
    list_per_page = 20              # default 20 ta
    list_max_show_all = 200         # "Show all" limiti
    list_per_page_options = [20, 50, 100]

@admin.register(Attempt)
class AttemptAdmin(ModelAdmin):  # <-- admin.ModelAdmin emas
    list_display = ("id", "user", "test", "status", "started_at", "finished_at", "score", "percent")
    list_filter = ("status", "test")
    search_fields = ("user__username", "test__title", "test__code")
    list_per_page = 20              # default 20 ta
    list_max_show_all = 200         # "Show all" limiti
    list_per_page_options = [20, 50, 100]

@admin.register(AttemptAnswer)
class AttemptAnswerAdmin(ModelAdmin):  # <-- admin.ModelAdmin emas
    list_display = (
        "id",
        "attempt_id",
        "attempt_user",
        "attempt_test",
        "question_short",
        "correct_badge",
        "earned_points",
    )
    list_filter = ("attempt__test", "attempt__status")
    list_per_page = 50              # default 50 ta
    list_max_show_all = 200         # "Show all" limiti
    list_per_page_options = [20, 50, 100]
    list_select_related = ("attempt", "attempt__user", "attempt__test", "question")
    search_fields = (
        "attempt__id",
        "attempt__user__username",
        "attempt__test__title",
        "question__text",
    )
    def attempt_id(self, obj):
        return obj.attempt.id
    attempt_id.short_description = "Attempt ID"

    def attempt_user(self, obj):
        return obj.attempt.user
    attempt_user.short_description = "User"

    def attempt_test(self, obj):
        return obj.attempt.test
    attempt_test.short_description = "Test"

    def question_short(self, obj):
        return obj.question.text[:80]
    question_short.short_description = "Question"

    def correct_badge(self, obj):
        if obj.is_correct:
            return mark_safe('<span style="color:#16a34a;font-weight:600;">✔ To‘g‘ri</span>')
        return mark_safe('<span style="color:#dc2626;font-weight:600;">✘ Noto‘g‘ri</span>')
    correct_badge.short_description = "Holat"

@admin.register(TestRule)
class TestRuleAdmin(ModelAdmin):
    list_display = ("id", "test", "branch", "department", "attempts_limit", "duration_minutes","deadline","is_active",)
    list_filter = ("test", "branch", "department",)
    list_display_links = ("test",)
    list_per_page = 20              # default 20 ta
    list_max_show_all = 200         # "Show all" limiti
    list_per_page_options = [20, 50, 100]


class AttemptPolicyForm(forms.ModelForm):
    class Meta:
        model = AttemptPolicy
        fields = "__all__"

    def clean(self):
        cleaned = super().clean()
        scope = cleaned.get("scope")
        branch = cleaned.get("branch")
        department = cleaned.get("department")
        user = cleaned.get("user")

        # Scope ga qarab faqat bittasi to‘ldirilishi kerak
        if scope == "branch" and not branch:
            raise forms.ValidationError("Branch tanlanishi kerak.")
        if scope == "department" and not department:
            raise forms.ValidationError("Department tanlanishi kerak.")
        if scope == "user" and not user:
            raise forms.ValidationError("User tanlanishi kerak.")

        return cleaned


@admin.register(AttemptPolicy)
class AttemptPolicyAdmin(ModelAdmin):
    form = AttemptPolicyForm

    list_display = (
        "id",
        "test",
        "scope_badge",
        "target_object",
        "extra_attempts",
        "updated_at",
    )

    list_filter = ("scope", "test")
    autocomplete_fields = ("test", "user", "branch", "department")

    fieldsets = (
        ("📘 Test ma’lumoti", {
            "fields": ("test",),
        }),
        ("🎯 Kim uchun?", {
            "fields": ("scope", "branch", "department", "user"),
        }),
        ("➕ Qo‘shimcha urinish", {
            "fields": ("extra_attempts",),
        }),
    )

    def scope_badge(self, obj):
        colors = {
            "branch": "#2563eb",
            "department": "#9333ea",
            "user": "#16a34a",
        }
        color = colors.get(obj.scope, "#64748b")
        return format_html(
            '<span style="background:{};color:white;padding:4px 10px;border-radius:12px;font-size:12px;">{}</span>',
            color,
            obj.scope.upper()
        )
    scope_badge.short_description = "Scope"

    def target_object(self, obj):
        return obj.user or obj.department or obj.branch
    target_object.short_description = "Obyekt"
